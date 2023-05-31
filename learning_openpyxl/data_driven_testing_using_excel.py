from openpyxl import Workbook

# Create a workbook object
wb = Workbook()

# A workbook is always created with a worksheet, get the active worksheet
ws = wb.active

# Create new worksheets
ws1 = wb.create_sheet("MySheet1")  # will be added at the end by default
ws2 = wb.create_sheet("MySheet2", 0)  # will be added at the first position
ws3 = wb.create_sheet("MySheet3", -1)  # will be added at the secondlastposition

# Change the title of the worksheet
ws.title = "New Title"

# Change the sheet colour
ws.sheet_properties.tabColor = "1072BA"


# Print all the worksheets
print(wb.sheetnames)

wb.save("MyFirstWorkBook.xls")
